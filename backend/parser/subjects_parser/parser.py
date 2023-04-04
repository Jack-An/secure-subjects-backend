import json

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class BaseParser:

    @classmethod
    def single_parser(cls, subject_rows):
        first_line = subject_rows[0]
        tag = first_line[1].value
        title = first_line[3].value
        answer = first_line[4].value
        options = []
        for option in subject_rows[1:]:
            option_val = option[3].value
            key = option_val[0]
            text = option_val[2:]
            options.append({"key": key, "option": text})
        subject = {
            "tag": tag,
            "title": title,
            "answer": answer,
            "options": options,
            "type": "single",
        }
        return subject

    @classmethod
    def multiple_parser(cls, subject_rows):
        first_line = subject_rows[0]
        tag = first_line[1].value
        title = first_line[3].value
        answer = first_line[4].value
        try:
            answer = [val for val in answer]
        except Exception:
            print(title)
            raise
        options = []
        for option in subject_rows[1:]:
            option_val = option[3].value
            key = option_val[0]
            text = option_val[2:]
            options.append({"key": key, "option": text})
        subject = {
            "tag": tag,
            "title": title,
            "answer": answer,
            "options": options,
            "type": "multiple",
        }
        return subject

    @classmethod
    def judge_parser(cls, subject_rows):
        first_line = subject_rows[0]
        tag = first_line[1].value
        title = first_line[3].value
        answer = first_line[4].value
        options = []
        for option in subject_rows[1:]:
            option_val = option[3].value
            key = option_val[0]
            text = option_val[2:]
            options.append({"key": key, "option": text})
        subject = {
            "tag": tag,
            "title": title,
            "answer": answer,
            "options": options,
            "type": "judge",
        }
        return subject



class SubjectParser:

    @classmethod
    def single_subject_paser(cls, rows):
        subjects = []
        for i in range(0, len(rows), 5):
            subject_rows = rows[i:i + 5]
            subject = BaseParser.single_parser(subject_rows)
            subjects.append(subject)
        return subjects

    @classmethod
    def multiple_subject_parser(cls, rows):
        subjects = []
        start_idx = 0
        for i in range(1, len(rows)):
            if rows[i][0].value:
                end_idx = i
                subjects.append(rows[start_idx:end_idx])
                start_idx = i
        subjects.append(rows[start_idx:])

        res = []
        for item in subjects:
            subject = BaseParser.multiple_parser(item)
            res.append(subject)
        return res

    @classmethod
    def judge_subject_parser(cls, rows):
        subjects = []
        for i in range(0, len(rows), 3):
            subject_rows = rows[i:i + 3]
            subject = BaseParser.judge_parser(subject_rows)
            subjects.append(subject)
        return subjects

    @classmethod
    def _split_sample_subjects(cls, rows):
        subjects = []
        start_idx = 0
        for i in range(1, len(rows)):
            if rows[i][0].value:
                end_idx = i
                subjects.append(rows[start_idx:end_idx])
                start_idx = i
        subjects.append(rows[start_idx:])
        return subjects

    @classmethod
    def _split_sample_sub_subjects(cls, rows):
        subjects = []
        start_idx = 0
        for i in range(1, len(rows)):
            if rows[i][2].value and rows[i][2].value.strip():
                end_idx = i
                subjects.append(rows[start_idx: end_idx])
                start_idx = i
        subjects.append(rows[start_idx:])
        return subjects

    @classmethod
    def sample_subject_parser(cls, rows):
        subjects = []
        _subjects = cls._split_sample_subjects(rows)
        for _item in _subjects:
            first_line = _item[0]
            tag = first_line[1].value
            title = first_line[3].value
            sub_rows = _item[1:]
            sub_subjects = cls._split_sample_sub_subjects(sub_rows)
            subs = []
            for _sub in sub_subjects:
                if len(_sub) == 3:
                    _sub_subject = BaseParser.judge_parser(_sub)
                elif len(_sub) == 5:
                    _sub_subject = BaseParser.single_parser(_sub)
                else:
                    _sub_subject = BaseParser.multiple_parser(_sub)
                subs.append(_sub_subject)

            subject = {
                "title": title,
                "tag": tag,
                "subs": subs,
                "type": "sample"
            }
            subjects.append(subject)
        return subjects

    @classmethod
    def clear_rows(cls, rows):
        valid_rows = []
        for k in rows:
            if k[3].value and k[3].value.strip():
                valid_rows.append(k)
        return valid_rows

    @classmethod
    def read_excel(cls):
        wb: Workbook = load_workbook(filename="../lib/建设工程安全生产管理.xlsx", data_only=True)
        sheet_key_map = {
            '单选': "single",
            "多选": "multiple",
            "判断": "judge",
            "案例": "sample"
        }
        result = {}
        for sheet_name in wb.sheetnames:
            subject_key = sheet_key_map[sheet_name]
            table = wb[sheet_name]
            rows = [row for row in table.rows][1:]
            valid_rows = cls.clear_rows(rows)
            if subject_key == "single":
                subjects = cls.single_subject_paser(valid_rows)
            elif subject_key == "multiple":
                subjects = cls.multiple_subject_parser(valid_rows)
            elif subject_key == "judge":
                subjects = cls.judge_subject_parser(valid_rows)
            else:
                subjects = cls.sample_subject_parser(valid_rows)

            with open(f'{subject_key}_管理.json', 'w') as f:
                json.dump(subjects, f)


if __name__ == '__main__':
    SubjectParser.read_excel()




